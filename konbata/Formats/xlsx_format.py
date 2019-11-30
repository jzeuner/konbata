"""
    Loader and Parser for the xlsx / xlsm files.

    This interface focusses on Excel 2010 xlsx file format.
    With the help of the openpyxl libary.

    https://openpyxl.readthedocs.io/en/stable/
"""

from konbata.Data.Data import DataNode, DataTree
import openpyxl
from konbata.Formats.Format import Format


def xlsx_toTree(file_name, options=None):
    """
    Function transforms a xlsx file into a DataTree.
    Therefore it needs the path of the original file.

    Parameters
    ----------
    file_name: str

    Returns
    -------
    tree: DataTree
    """

    # TODO: add options guess_type, data_only, keep_vba

    xlsx_wb = openpyxl.load_workbook(filename=file_name)

    tree = DataTree(tree_type='xlsx')

    # For each sheet create a node
    for sheet_name in xlsx_wb.sheetnames:
        sheet_node = DataNode('Sheet:' + sheet_name)
        sheet = xlsx_wb[sheet_name]
        # go through each cell in matrix max_rows x max_cols
        # and create tree like csv file

        i = 0
        for row in sheet.rows:
            row_node = DataNode('Row%s' % i)
            i += 1
            for cell in row:
                col_node = DataNode(str(cell.value))
                row_node.add(col_node)
            sheet_node.add(row_node)
        tree.root.add(sheet_node)

    return tree


def xlsx_fromTree(tree, file_name, options=None):
    """
    Function transforms a DataTree into a xlsx file.

    Parameters
    ----------
    tree: DataTree
    file_name: str
    """

    if not isinstance(tree, DataTree):
        raise TypeError('tree must be type of DataTree and not ', type(tree))

    if tree.tree_type != 'xlsx' or tree.height() != 4:
        # Height of tree needs to be flatten or need to be increased
        if tree.height() > 4:
            tree.minimize_height(tree.height()-4)
        elif tree.height() < 4:
            tree.increase_height(4-tree.height())

    # Here we have a tree of the right shape
    xlsx_wb = openpyxl.Workbook()

    for sheet_node in tree.root.children:
        sheet = xlsx_wb.create_sheet(title=sheet_node.data)
        row_i = 1
        for row_node in sheet_node.children:
            col_j = 1
            for col_node in row_node.children:
                sheet.cell(column=col_j, row=row_i, value=col_node.data)
                col_j += 1
            row_i += 1

    xlsx_wb.save(filename=file_name)


xlsx_format = Format('xlsx', [';', ','], xlsx_toTree, xlsx_fromTree,
                     pre_open=False)
